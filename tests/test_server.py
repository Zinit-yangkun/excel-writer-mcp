"""Tests for excel_writer_mcp.server — covers all tools and helpers."""

import shutil
from pathlib import Path

import openpyxl
import pytest

from excel_writer_mcp import server as _srv

# Helpers are plain functions
_is_xlsm = _srv._is_xlsm
_load_workbook = _srv._load_workbook
_save_workbook = _srv._save_workbook
_get_ws = _srv._get_ws

# Tools are FunctionTool wrappers — unwrap via .fn
copy_file = _srv.copy_file.fn
create_workbook = _srv.create_workbook.fn
format_cells = _srv.format_cells.fn
get_workbook_info = _srv.get_workbook_info.fn
manage_sheets = _srv.manage_sheets.fn
merge_cells = _srv.merge_cells.fn
modify_rows_columns = _srv.modify_rows_columns.fn
read_data = _srv.read_data.fn
write_cells = _srv.write_cells.fn
write_csv = _srv.write_csv.fn
write_data = _srv.write_data.fn
read_csv = _srv.read_csv.fn


@pytest.fixture()
def tmp_dir(tmp_path):
    """Provide a temporary directory path as string."""
    return str(tmp_path)


@pytest.fixture()
def xlsx_path(tmp_dir):
    """Create a basic .xlsx workbook and return its path."""
    p = f"{tmp_dir}/test.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    wb.save(p)
    wb.close()
    return p


@pytest.fixture()
def xlsx_with_data(tmp_dir):
    """Create a .xlsx with sample data."""
    p = f"{tmp_dir}/data.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Age", "Score"])
    ws.append(["Alice", 30, 90])
    ws.append(["Bob", 25, 85])
    wb.save(p)
    wb.close()
    return p


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

class TestHelpers:
    def test_is_xlsm(self):
        assert _is_xlsm("foo.xlsm") is True
        assert _is_xlsm("foo.XLSM") is True
        assert _is_xlsm("foo.xlsx") is False

    def test_load_workbook_not_found(self):
        with pytest.raises(FileNotFoundError):
            _load_workbook("/nonexistent/file.xlsx")

    def test_load_workbook_xls_rejected(self, tmp_dir):
        p = f"{tmp_dir}/old.xls"
        Path(p).touch()
        with pytest.raises(ValueError, match=r"\.xls format is not supported"):
            _load_workbook(p)

    def test_load_workbook_xlsx(self, xlsx_path):
        wb = _load_workbook(xlsx_path)
        assert "Sheet1" in wb.sheetnames
        wb.close()

    def test_save_workbook_creates_parent_dirs(self, tmp_dir):
        p = f"{tmp_dir}/a/b/c/test.xlsx"
        wb = openpyxl.Workbook()
        _save_workbook(wb, p)
        wb.close()
        assert Path(p).exists()

    def test_get_ws_returns_named_sheet(self, xlsx_path):
        wb = _load_workbook(xlsx_path)
        ws = _get_ws(wb, "Sheet1")
        assert ws.title == "Sheet1"
        wb.close()

    def test_get_ws_raises_when_name_not_found(self, xlsx_path):
        wb = _load_workbook(xlsx_path)
        with pytest.raises(KeyError, match="NonExistent"):
            _get_ws(wb, "NonExistent")
        wb.close()

    def test_get_ws_returns_active_when_none(self, xlsx_path):
        wb = _load_workbook(xlsx_path)
        ws = _get_ws(wb, None)
        assert ws == wb.active
        wb.close()


# ---------------------------------------------------------------------------
# create_workbook
# ---------------------------------------------------------------------------

class TestCreateWorkbook:
    def test_create_xlsx(self, tmp_dir):
        p = f"{tmp_dir}/new.xlsx"
        result = create_workbook(p, sheet_name="MySheet")
        assert "Created workbook" in result
        wb = openpyxl.load_workbook(p)
        assert "MySheet" in wb.sheetnames
        wb.close()

    def test_create_xlsm_rejected(self, tmp_dir):
        p = f"{tmp_dir}/new.xlsm"
        with pytest.raises(ValueError, match="Cannot create .xlsm from scratch"):
            create_workbook(p)

    def test_auto_add_xlsx_extension(self, tmp_dir):
        p = f"{tmp_dir}/noext"
        result = create_workbook(p)
        assert "Created workbook" in result
        assert Path(f"{tmp_dir}/noext.xlsx").exists()

    def test_default_sheet_name(self, tmp_dir):
        p = f"{tmp_dir}/default.xlsx"
        create_workbook(p)
        wb = openpyxl.load_workbook(p)
        assert "Sheet1" in wb.sheetnames
        wb.close()


# ---------------------------------------------------------------------------
# copy_file
# ---------------------------------------------------------------------------

class TestCopyFile:
    def test_copy_success(self, xlsx_path, tmp_dir):
        dest = f"{tmp_dir}/copy.xlsx"
        result = copy_file(xlsx_path, dest)
        assert "Copied" in result
        assert Path(dest).exists()

    def test_copy_source_not_found(self, tmp_dir):
        with pytest.raises(FileNotFoundError, match="Source file not found"):
            copy_file(f"{tmp_dir}/nope.xlsx", f"{tmp_dir}/dest.xlsx")

    def test_copy_dest_exists(self, xlsx_path, tmp_dir):
        dest = f"{tmp_dir}/dup.xlsx"
        Path(dest).touch()
        with pytest.raises(FileExistsError, match="Will not overwrite"):
            copy_file(xlsx_path, dest)

    def test_copy_creates_parent_dirs(self, xlsx_path, tmp_dir):
        dest = f"{tmp_dir}/x/y/z/copy.xlsx"
        result = copy_file(xlsx_path, dest)
        assert "Copied" in result


# ---------------------------------------------------------------------------
# get_workbook_info
# ---------------------------------------------------------------------------

class TestGetWorkbookInfo:
    def test_basic_info(self, xlsx_path):
        info = get_workbook_info(xlsx_path)
        assert info["path"] == xlsx_path
        assert info["is_xlsm"] is False
        assert info["has_vba"] is False
        assert len(info["sheets"]) == 1
        assert info["sheets"][0]["name"] == "Sheet1"

    def test_info_with_data(self, xlsx_with_data):
        info = get_workbook_info(xlsx_with_data)
        sheet = info["sheets"][0]
        assert sheet["name"] == "Data"
        assert sheet["max_row"] == 3
        assert sheet["max_column"] == 3


# ---------------------------------------------------------------------------
# manage_sheets
# ---------------------------------------------------------------------------

class TestManageSheets:
    def test_create_sheet(self, xlsx_path):
        result = manage_sheets(xlsx_path, "create", "NewSheet")
        assert "create" in result
        wb = openpyxl.load_workbook(xlsx_path)
        assert "NewSheet" in wb.sheetnames
        wb.close()

    def test_create_sheet_with_position(self, xlsx_path):
        manage_sheets(xlsx_path, "create", "First", position=0)
        wb = openpyxl.load_workbook(xlsx_path)
        assert wb.sheetnames[0] == "First"
        wb.close()

    def test_create_duplicate_sheet(self, xlsx_path):
        with pytest.raises(ValueError, match="already exists"):
            manage_sheets(xlsx_path, "create", "Sheet1")

    def test_delete_sheet(self, xlsx_path):
        manage_sheets(xlsx_path, "create", "ToDelete")
        result = manage_sheets(xlsx_path, "delete", "ToDelete")
        assert "delete" in result
        wb = openpyxl.load_workbook(xlsx_path)
        assert "ToDelete" not in wb.sheetnames
        wb.close()

    def test_delete_nonexistent_sheet(self, xlsx_path):
        with pytest.raises(KeyError, match="not found"):
            manage_sheets(xlsx_path, "delete", "NoSuch")

    def test_delete_only_sheet(self, xlsx_path):
        with pytest.raises(ValueError, match="only sheet"):
            manage_sheets(xlsx_path, "delete", "Sheet1")

    def test_rename_sheet(self, xlsx_path):
        result = manage_sheets(xlsx_path, "rename", "Sheet1", new_name="Renamed")
        assert "rename" in result
        wb = openpyxl.load_workbook(xlsx_path)
        assert "Renamed" in wb.sheetnames
        wb.close()

    def test_rename_nonexistent_sheet(self, xlsx_path):
        with pytest.raises(KeyError, match="not found"):
            manage_sheets(xlsx_path, "rename", "NoSuch", new_name="X")

    def test_rename_without_new_name(self, xlsx_path):
        with pytest.raises(ValueError, match="new_name is required"):
            manage_sheets(xlsx_path, "rename", "Sheet1")

    def test_unknown_action(self, xlsx_path):
        with pytest.raises(ValueError, match="Unknown action"):
            manage_sheets(xlsx_path, "unknown", "Sheet1")


# ---------------------------------------------------------------------------
# read_data
# ---------------------------------------------------------------------------

class TestReadData:
    def test_read_all(self, xlsx_with_data):
        result = read_data(xlsx_with_data, sheet_name="Data")
        assert result["total_rows"] == 3
        assert result["data"][0] == ["Name", "Age", "Score"]
        assert result["data"][1] == ["Alice", "30", "90"]

    def test_read_subrange(self, xlsx_with_data):
        result = read_data(xlsx_with_data, start_row=2, start_col=1, end_row=2, end_col=2)
        assert result["total_rows"] == 1
        assert result["data"][0] == ["Alice", "30"]

    def test_read_includes_merged_cells(self, xlsx_path):
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        ws.merge_cells("A1:B1")
        wb.save(xlsx_path)
        wb.close()

        result = read_data(xlsx_path, include_merged_cells=True)
        assert "merged_cells" in result
        assert "A1:B1" in result["merged_cells"]

    def test_read_defaults_to_active_sheet(self, xlsx_with_data):
        result = read_data(xlsx_with_data)
        assert result["sheet"] == "Data"


# ---------------------------------------------------------------------------
# write_data
# ---------------------------------------------------------------------------

class TestWriteData:
    def test_write_and_read_back(self, xlsx_path):
        write_data(xlsx_path, data=[["X", "Y"], [1, 2]])
        result = read_data(xlsx_path)
        assert result["data"][0] == ["X", "Y"]
        assert result["data"][1] == ["1", "2"]

    def test_write_creates_new_xlsx(self, tmp_dir):
        p = f"{tmp_dir}/auto.xlsx"
        result = write_data(p, data=[["Hello"]])
        assert "Wrote" in result
        assert Path(p).exists()

    def test_write_xlsm_must_exist(self, tmp_dir):
        p = f"{tmp_dir}/no.xlsm"
        with pytest.raises(ValueError, match=".xlsm file must exist"):
            write_data(p, data=[["A"]])

    def test_write_auto_fit_width(self, xlsx_path):
        result = write_data(xlsx_path, data=[["LongColumnHeader", "Short"]], auto_fit_width=True)
        assert "Wrote" in result
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        assert ws.column_dimensions["A"].width > 0
        wb.close()

    def test_write_with_formula(self, xlsx_path):
        write_data(xlsx_path, data=[[10], [20], ["=SUM(A1:A2)"]])
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        assert ws["A3"].value == "=SUM(A1:A2)"
        wb.close()

    def test_write_auto_add_extension(self, tmp_dir):
        p = f"{tmp_dir}/noext"
        result = write_data(p, data=[["a"]])
        assert "Wrote" in result
        assert Path(f"{tmp_dir}/noext.xlsx").exists()

    def test_write_with_offset(self, xlsx_path):
        write_data(xlsx_path, data=[["val"]], start_row=3, start_col=2)
        wb = openpyxl.load_workbook(xlsx_path)
        assert wb.active["B3"].value == "val"
        wb.close()

    def test_write_empty_data(self, xlsx_path):
        result = write_data(xlsx_path, data=[])
        assert "Wrote 0 rows" in result


# ---------------------------------------------------------------------------
# write_cells
# ---------------------------------------------------------------------------

class TestWriteCells:
    def test_write_specific_cells(self, xlsx_path):
        result = write_cells(xlsx_path, cells={"A1": "Hello", "C3": 42})
        assert "Wrote 2 cell(s)" in result
        wb = openpyxl.load_workbook(xlsx_path)
        assert wb.active["A1"].value == "Hello"
        assert wb.active["C3"].value == 42
        wb.close()

    def test_write_cells_file_not_found(self, tmp_dir):
        with pytest.raises(FileNotFoundError, match="File not found"):
            write_cells(f"{tmp_dir}/nope.xlsx", cells={"A1": 1})

    def test_write_cells_with_formula(self, xlsx_path):
        write_cells(xlsx_path, cells={"A1": 10, "A2": "=A1*2"})
        wb = openpyxl.load_workbook(xlsx_path)
        assert wb.active["A2"].value == "=A1*2"
        wb.close()


# ---------------------------------------------------------------------------
# modify_rows_columns
# ---------------------------------------------------------------------------

class TestModifyRowsColumns:
    def test_insert_rows(self, xlsx_with_data):
        result = modify_rows_columns(xlsx_with_data, "insert_rows", 2, count=2)
        assert "insert_rows" in result
        wb = openpyxl.load_workbook(xlsx_with_data)
        # Original row 2 (Alice) should now be at row 4
        assert wb.active.cell(4, 1).value == "Alice"
        wb.close()

    def test_delete_rows(self, xlsx_with_data):
        modify_rows_columns(xlsx_with_data, "delete_rows", 2)
        wb = openpyxl.load_workbook(xlsx_with_data)
        # Row 2 should now be Bob
        assert wb.active.cell(2, 1).value == "Bob"
        wb.close()

    def test_insert_cols(self, xlsx_with_data):
        modify_rows_columns(xlsx_with_data, "insert_cols", 1)
        wb = openpyxl.load_workbook(xlsx_with_data)
        # Name should now be in column B
        assert wb.active.cell(1, 2).value == "Name"
        wb.close()

    def test_delete_cols(self, xlsx_with_data):
        modify_rows_columns(xlsx_with_data, "delete_cols", 1)
        wb = openpyxl.load_workbook(xlsx_with_data)
        # First column should now be Age
        assert wb.active.cell(1, 1).value == "Age"
        wb.close()

    def test_unknown_action(self, xlsx_with_data):
        with pytest.raises(ValueError, match="Unknown action"):
            modify_rows_columns(xlsx_with_data, "bad_action", 1)


# ---------------------------------------------------------------------------
# merge_cells
# ---------------------------------------------------------------------------

class TestMergeCells:
    def test_merge(self, xlsx_path):
        result = merge_cells(xlsx_path, "A1:C1")
        assert "Merged" in result
        wb = openpyxl.load_workbook(xlsx_path)
        merged = [str(r) for r in wb.active.merged_cells.ranges]
        assert "A1:C1" in merged
        wb.close()

    def test_unmerge(self, xlsx_path):
        merge_cells(xlsx_path, "A1:C1")
        result = merge_cells(xlsx_path, "A1:C1", unmerge=True)
        assert "Unmerged" in result
        wb = openpyxl.load_workbook(xlsx_path)
        assert len(wb.active.merged_cells.ranges) == 0
        wb.close()


# ---------------------------------------------------------------------------
# format_cells
# ---------------------------------------------------------------------------

class TestFormatCells:
    def test_bold_and_italic(self, xlsx_with_data):
        result = format_cells(xlsx_with_data, "A1:C1", bold=True, italic=True)
        assert "Formatted" in result
        wb = openpyxl.load_workbook(xlsx_with_data)
        cell = wb.active["A1"]
        assert cell.font.bold is True
        assert cell.font.italic is True
        wb.close()

    def test_font_size_and_color(self, xlsx_with_data):
        format_cells(xlsx_with_data, "A1", font_size=16, font_color="FF0000")
        wb = openpyxl.load_workbook(xlsx_with_data)
        cell = wb.active["A1"]
        assert cell.font.size == 16
        wb.close()

    def test_bg_color(self, xlsx_with_data):
        format_cells(xlsx_with_data, "A1", bg_color="00FF00")
        wb = openpyxl.load_workbook(xlsx_with_data)
        assert wb.active["A1"].fill.start_color.rgb == "0000FF00"
        wb.close()

    def test_alignment(self, xlsx_with_data):
        format_cells(xlsx_with_data, "A1", horizontal="center", vertical="top", wrap_text=True)
        wb = openpyxl.load_workbook(xlsx_with_data)
        cell = wb.active["A1"]
        assert cell.alignment.horizontal == "center"
        assert cell.alignment.vertical == "top"
        assert cell.alignment.wrap_text is True
        wb.close()

    def test_border(self, xlsx_with_data):
        format_cells(xlsx_with_data, "A1", border_style="thin")
        wb = openpyxl.load_workbook(xlsx_with_data)
        cell = wb.active["A1"]
        assert cell.border.left.style == "thin"
        wb.close()

    def test_number_format(self, xlsx_with_data):
        format_cells(xlsx_with_data, "B2", number_format="#,##0.00")
        wb = openpyxl.load_workbook(xlsx_with_data)
        assert wb.active["B2"].number_format == "#,##0.00"
        wb.close()

    def test_column_width_and_row_height(self, xlsx_with_data):
        format_cells(xlsx_with_data, "A1:B2", column_width=20, row_height=30)
        wb = openpyxl.load_workbook(xlsx_with_data)
        ws = wb.active
        assert ws.column_dimensions["A"].width == 20
        assert ws.column_dimensions["B"].width == 20
        assert ws.row_dimensions[1].height == 30
        assert ws.row_dimensions[2].height == 30
        wb.close()


# ---------------------------------------------------------------------------
# CSV read/write
# ---------------------------------------------------------------------------

class TestCsv:
    def test_write_and_read_roundtrip(self, tmp_dir):
        p = f"{tmp_dir}/data.csv"
        data = [["Name", "Age"], ["Alice", 30], ["Bob", 25]]
        result = write_csv(p, data)
        assert "Wrote 3 rows" in result
        out = read_csv(p)
        assert out["total_rows"] == 3
        assert out["returned_rows"] == 3
        assert out["has_more"] is False
        assert out["data"][0] == ["Name", "Age"]
        assert out["data"][1] == ["Alice", 30]
        assert out["data"][2] == ["Bob", 25]

    def test_append_mode(self, tmp_dir):
        p = f"{tmp_dir}/append.csv"
        write_csv(p, [["a", "b"]])
        result = write_csv(p, [["c", "d"]], append=True)
        assert "Appended 1 rows" in result
        out = read_csv(p)
        assert out["total_rows"] == 2
        assert out["data"][0] == ["a", "b"]
        assert out["data"][1] == ["c", "d"]

    def test_chunked_read(self, tmp_dir):
        p = f"{tmp_dir}/chunked.csv"
        rows = [[f"r{i}"] for i in range(5)]
        write_csv(p, rows)

        # Read first 2 rows
        out = read_csv(p, max_rows=2)
        assert out["total_rows"] == 5
        assert out["returned_rows"] == 2
        assert out["has_more"] is True
        assert out["next_start_row"] == 3
        assert out["data"] == [["r0"], ["r1"]]

        # Read next chunk
        out2 = read_csv(p, start_row=out["next_start_row"], max_rows=2)
        assert out2["returned_rows"] == 2
        assert out2["has_more"] is True
        assert out2["data"] == [["r2"], ["r3"]]

        # Read last chunk
        out3 = read_csv(p, start_row=out2["next_start_row"], max_rows=2)
        assert out3["returned_rows"] == 1
        assert out3["has_more"] is False
        assert "next_start_row" not in out3
        assert out3["data"] == [["r4"]]

    def test_read_file_not_found(self, tmp_dir):
        with pytest.raises(FileNotFoundError):
            read_csv(f"{tmp_dir}/nope.csv")

    def test_numeric_conversion(self, tmp_dir):
        p = f"{tmp_dir}/nums.csv"
        write_csv(p, [["10", "3.14", "", "hello"]])
        out = read_csv(p)
        row = out["data"][0]
        assert row[0] == 10
        assert row[1] == 3.14
        assert row[2] is None
        assert row[3] == "hello"

    def test_empty_file(self, tmp_dir):
        p = f"{tmp_dir}/empty.csv"
        write_csv(p, [])
        out = read_csv(p)
        assert out["total_rows"] == 0
        assert out["data"] == []
        assert out["has_more"] is False

    def test_write_creates_parent_dirs(self, tmp_dir):
        p = f"{tmp_dir}/a/b/c/nested.csv"
        result = write_csv(p, [["x"]])
        assert "Wrote" in result
        assert Path(p).exists()
