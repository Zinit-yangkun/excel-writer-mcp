"""Excel Writer MCP Server - supports .xlsx and .xlsm (with VBA macro preservation)."""

import logging
import shutil
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from fastmcp import FastMCP

logger = logging.getLogger(__name__)

mcp = FastMCP("excel-writer-mcp")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _is_xlsm(path: str) -> bool:
    return Path(path).suffix.lower() == ".xlsm"


def _load_workbook(path: str) -> openpyxl.Workbook:
    """Load workbook with keep_vba=True for .xlsm files."""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"File not found: {path}")
    if p.suffix.lower() == ".xls":
        raise ValueError(
            f"The .xls format is not supported. "
            f"Please convert '{p.name}' to .xlsx first (e.g. open in Excel and Save As .xlsx)."
        )
    return openpyxl.load_workbook(path, keep_vba=_is_xlsm(path))


def _save_workbook(wb: openpyxl.Workbook, path: str) -> None:
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _get_ws(wb: openpyxl.Workbook, sheet_name: str | None):
    if sheet_name and sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return wb.active


# ---------------------------------------------------------------------------
# Tools
# ---------------------------------------------------------------------------

@mcp.tool()
def create_workbook(path: str, sheet_name: str = "Sheet1") -> str:
    """Create a new empty .xlsx workbook. For .xlsm, use copy_file from an existing template instead.

    Args:
        path: File path for the new workbook (.xlsx).
        sheet_name: Name of the initial worksheet.
    """
    p = Path(path)
    if p.suffix.lower() not in (".xlsx", ".xlsm"):
        path = str(p.with_suffix(".xlsx"))
    if _is_xlsm(path):
        return "Error: cannot create .xlsm from scratch. Use copy_file from an existing .xlsm template."
    wb = openpyxl.Workbook()
    wb.active.title = sheet_name
    _save_workbook(wb, path)
    wb.close()
    return f"Created workbook: {path} with sheet '{sheet_name}'"


@mcp.tool()
def copy_file(source_path: str, dest_path: str) -> str:
    """Safely copy any file to a new path. Will NOT overwrite an existing file.
    Preserves VBA macros when copying .xlsm files.

    Args:
        source_path: Path of the source file.
        dest_path: Path for the destination file.
    """
    if not Path(source_path).exists():
        return f"Error: source file not found: {source_path}"
    if Path(dest_path).exists():
        return f"Error: destination already exists: {dest_path}. Will not overwrite."
    Path(dest_path).parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_path, dest_path)
    return f"Copied '{source_path}' -> '{dest_path}'"


@mcp.tool()
def get_workbook_info(path: str) -> dict[str, Any]:
    """Get workbook metadata: sheet names, dimensions, file type, and VBA status.

    Args:
        path: Path to the workbook (.xlsx or .xlsm).
    """
    wb = _load_workbook(path)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        sheets.append({
            "name": name,
            "dimensions": ws.dimensions,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
        })
    has_vba = wb.vba_archive is not None if hasattr(wb, "vba_archive") else False
    vba_files = wb.vba_archive.namelist() if has_vba and wb.vba_archive else []
    wb.close()
    return {
        "path": path,
        "is_xlsm": _is_xlsm(path),
        "has_vba": has_vba,
        "vba_archive_files": vba_files,
        "sheets": sheets,
    }


@mcp.tool()
def manage_sheets(
    path: str,
    action: str,
    sheet_name: str,
    new_name: str | None = None,
    position: int | None = None,
) -> str:
    """Create, delete, or rename a worksheet.

    Args:
        path: Path to the workbook.
        action: One of "create", "delete", "rename".
        sheet_name: Target sheet name (for delete/rename: existing name; for create: new name).
        new_name: New name when action is "rename".
        position: Optional index (0-based) when action is "create".
    """
    wb = _load_workbook(path)
    if action == "create":
        if sheet_name in wb.sheetnames:
            wb.close()
            return f"Error: sheet '{sheet_name}' already exists."
        wb.create_sheet(title=sheet_name, index=position)
    elif action == "delete":
        if sheet_name not in wb.sheetnames:
            wb.close()
            return f"Error: sheet '{sheet_name}' not found."
        if len(wb.sheetnames) == 1:
            wb.close()
            return "Error: cannot delete the only sheet."
        del wb[sheet_name]
    elif action == "rename":
        if sheet_name not in wb.sheetnames:
            wb.close()
            return f"Error: sheet '{sheet_name}' not found."
        if not new_name:
            wb.close()
            return "Error: new_name is required for rename."
        wb[sheet_name].title = new_name
    else:
        wb.close()
        return f"Error: unknown action '{action}'. Use: create, delete, rename."
    _save_workbook(wb, path)
    wb.close()
    return f"Sheet '{sheet_name}' {action}d in {path}" + (f" -> '{new_name}'" if new_name else "")


@mcp.tool()
def read_data(
    path: str,
    sheet_name: str | None = None,
    start_row: int = 1,
    start_col: int = 1,
    end_row: int | None = None,
    end_col: int | None = None,
    include_merged_cells: bool = False,
) -> dict[str, Any]:
    """Read data from a worksheet range.

    Args:
        path: Path to the workbook.
        sheet_name: Sheet name (defaults to active sheet).
        start_row: First row (1-based).
        start_col: First column (1-based).
        end_row: Last row (defaults to max row).
        end_col: Last column (defaults to max column).
        include_merged_cells: If true, include merged cell ranges in the response. Enable this before using write_cells on sheets with merged cells.
    """
    wb = _load_workbook(path)
    ws = _get_ws(wb, sheet_name)
    end_row = end_row or ws.max_row
    end_col = end_col or ws.max_column
    rows = []
    for row in ws.iter_rows(min_row=start_row, max_row=end_row,
                            min_col=start_col, max_col=end_col, values_only=True):
        rows.append([str(c) if c is not None else None for c in row])
    result: dict[str, Any] = {
        "sheet": ws.title,
        "range": f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}",
        "total_rows": len(rows),
        "data": rows,
    }
    if include_merged_cells:
        result["merged_cells"] = [str(m) for m in ws.merged_cells.ranges]
    wb.close()
    return result


@mcp.tool()
def write_data(
    path: str,
    sheet_name: str | None = None,
    start_row: int = 1,
    start_col: int = 1,
    data: list[list[Any]] = [],
    auto_fit_width: bool = False,
) -> str:
    """Write a 2D array of data to a worksheet. Works with both .xlsx and .xlsm
    (VBA macros are automatically preserved for .xlsm).
    Also supports formulas — just include "=SUM(...)" as a cell value.

    Args:
        path: Path to the workbook. Created automatically for .xlsx if not exists.
        sheet_name: Target sheet (defaults to active sheet).
        start_row: Starting row (1-based).
        start_col: Starting column (1-based).
        data: 2D list of values, e.g. [["Name","Age"],["Alice",30],["Total","=SUM(B2:B2)"]].
        auto_fit_width: Auto-adjust column widths to fit content.
    """
    p = Path(path)
    if p.suffix.lower() not in (".xlsx", ".xlsm"):
        path = str(p.with_suffix(".xlsx"))
    if not Path(path).exists():
        if _is_xlsm(path):
            return "Error: .xlsm file must exist before writing. Use copy_file from a template."
        wb = openpyxl.Workbook()
    else:
        wb = _load_workbook(path)
    ws = _get_ws(wb, sheet_name)
    for r_idx, row in enumerate(data):
        for c_idx, value in enumerate(row):
            ws.cell(row=start_row + r_idx, column=start_col + c_idx, value=value)
    if auto_fit_width and data:
        for c_idx in range(len(data[0])):
            col_letter = get_column_letter(start_col + c_idx)
            max_len = max(
                (len(str(row[c_idx])) for row in data if c_idx < len(row) and row[c_idx] is not None),
                default=0,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 2, 60)
    _save_workbook(wb, path)
    wb.close()
    end_row = start_row + len(data) - 1
    end_col = start_col + (max(len(r) for r in data) - 1 if data else 0)
    rng = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    return f"Wrote {len(data)} rows to '{ws.title}' range {rng} in {path}"


@mcp.tool()
def write_cells(
    path: str,
    cells: dict[str, Any],
    sheet_name: str | None = None,
) -> str:
    """Write values to multiple specific cells by address. Ideal for worksheets
    with merged cells or non-contiguous writes. Also supports formulas.

    Args:
        path: Path to the workbook (.xlsx or .xlsm).
        cells: Mapping of cell address to value, e.g. {"A1": "Name", "B3": 100, "C5": "=SUM(C1:C4)"}.
        sheet_name: Target sheet (defaults to active sheet).
    """
    if not Path(path).exists():
        return f"Error: file not found: {path}"
    wb = _load_workbook(path)
    ws = _get_ws(wb, sheet_name)
    for cell_addr, value in cells.items():
        ws[cell_addr] = value
    _save_workbook(wb, path)
    wb.close()
    return f"Wrote {len(cells)} cell(s) to '{ws.title}' ({path}): {', '.join(cells.keys())}"


@mcp.tool()
def modify_rows_columns(
    path: str,
    action: str,
    index: int,
    count: int = 1,
    sheet_name: str | None = None,
) -> str:
    """Insert or delete rows/columns.

    Args:
        path: Path to the workbook.
        action: One of "insert_rows", "delete_rows", "insert_cols", "delete_cols".
        index: Row or column number (1-based) at which to insert/delete.
        count: Number of rows or columns to insert/delete.
        sheet_name: Target sheet (defaults to active sheet).
    """
    wb = _load_workbook(path)
    ws = _get_ws(wb, sheet_name)
    actions = {
        "insert_rows": lambda: ws.insert_rows(index, count),
        "delete_rows": lambda: ws.delete_rows(index, count),
        "insert_cols": lambda: ws.insert_cols(index, count),
        "delete_cols": lambda: ws.delete_cols(index, count),
    }
    fn = actions.get(action)
    if not fn:
        wb.close()
        return f"Error: unknown action '{action}'. Use: insert_rows, delete_rows, insert_cols, delete_cols."
    fn()
    _save_workbook(wb, path)
    wb.close()
    return f"{action} {count} at index {index} in '{ws.title}' ({path})"


@mcp.tool()
def merge_cells(path: str, range_string: str, unmerge: bool = False, sheet_name: str | None = None) -> str:
    """Merge or unmerge a range of cells.

    Args:
        path: Path to the workbook.
        range_string: Cell range like "A1:D1".
        unmerge: Set to true to unmerge instead of merge.
        sheet_name: Target sheet (defaults to active sheet).
    """
    wb = _load_workbook(path)
    ws = _get_ws(wb, sheet_name)
    if unmerge:
        ws.unmerge_cells(range_string)
    else:
        ws.merge_cells(range_string)
    _save_workbook(wb, path)
    wb.close()
    action = "Unmerged" if unmerge else "Merged"
    return f"{action} {range_string} in '{ws.title}' ({path})"


@mcp.tool()
def format_cells(
    path: str,
    range_string: str,
    sheet_name: str | None = None,
    bold: bool | None = None,
    italic: bool | None = None,
    font_size: int | None = None,
    font_color: str | None = None,
    bg_color: str | None = None,
    number_format: str | None = None,
    horizontal: str | None = None,
    vertical: str | None = None,
    wrap_text: bool | None = None,
    border_style: str | None = None,
    column_width: float | None = None,
    row_height: float | None = None,
) -> str:
    """Apply formatting to a range of cells. Works with both .xlsx and .xlsm.

    Args:
        path: Path to the workbook.
        range_string: Cell range like "A1:D1" or single cell "A1".
        sheet_name: Target sheet (defaults to active sheet).
        bold: Set font bold.
        italic: Set font italic.
        font_size: Font size in points.
        font_color: Font color as hex (e.g. "FF0000" for red).
        bg_color: Background fill color as hex.
        number_format: Excel number format (e.g. "#,##0.00", "yyyy-mm-dd").
        horizontal: Horizontal alignment ("left", "center", "right").
        vertical: Vertical alignment ("top", "center", "bottom").
        wrap_text: Enable text wrapping.
        border_style: Border style ("thin", "medium", "thick", "double").
        column_width: Set width for all columns in the range.
        row_height: Set height for all rows in the range.
    """
    wb = _load_workbook(path)
    ws = _get_ws(wb, sheet_name)

    font_kwargs: dict[str, Any] = {}
    if bold is not None:
        font_kwargs["bold"] = bold
    if italic is not None:
        font_kwargs["italic"] = italic
    if font_size is not None:
        font_kwargs["size"] = font_size
    if font_color is not None:
        font_kwargs["color"] = font_color

    fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid") if bg_color else None

    align_kwargs: dict[str, Any] = {}
    if horizontal:
        align_kwargs["horizontal"] = horizontal
    if vertical:
        align_kwargs["vertical"] = vertical
    if wrap_text is not None:
        align_kwargs["wrap_text"] = wrap_text
    alignment = Alignment(**align_kwargs) if align_kwargs else None

    border = None
    if border_style:
        side = Side(style=border_style)
        border = Border(left=side, right=side, top=side, bottom=side)

    target = ws[range_string]
    if hasattr(target, "value"):
        # Single cell (e.g. "A1")
        rows = ((target,),)
    elif isinstance(target, tuple) and target and not isinstance(target[0], tuple):
        # Single row (e.g. "A1:C1")
        rows = (target,)
    else:
        rows = target
    for row in rows:
        cells = row if isinstance(row, tuple) else (row,)
        for cell in cells:
            if font_kwargs:
                cell.font = Font(**{**{k: getattr(cell.font, k) for k in (
                    "name", "size", "bold", "italic", "color", "underline", "strikethrough"
                )}, **font_kwargs})
            if fill:
                cell.fill = fill
            if alignment:
                cell.alignment = alignment
            if border:
                cell.border = border
            if number_format:
                cell.number_format = number_format

    # Apply column width / row height for the range
    if column_width is not None or row_height is not None:
        min_col, min_row, max_col, max_row = range_boundaries(range_string)
        if column_width is not None:
            for c in range(min_col, (max_col or min_col) + 1):
                ws.column_dimensions[get_column_letter(c)].width = column_width
        if row_height is not None:
            for r in range(min_row, (max_row or min_row) + 1):
                ws.row_dimensions[r].height = row_height

    _save_workbook(wb, path)
    wb.close()
    return f"Formatted {range_string} in '{ws.title}' ({path})"


@mcp.tool()
def create_chart(
    path: str,
    chart_type: str,
    data_range: str,
    target_cell: str = "E1",
    sheet_name: str | None = None,
    title: str = "",
    categories_range: str | None = None,
    width: float = 15,
    height: float = 10,
) -> str:
    """Create a chart (bar, line, or pie) in the worksheet.

    Args:
        path: Path to the workbook.
        chart_type: One of "bar", "line", "pie".
        data_range: Data range for chart values, e.g. "B1:B10".
        target_cell: Cell where chart will be placed (default "E1").
        sheet_name: Target sheet (defaults to active sheet).
        title: Chart title.
        categories_range: Range for category labels, e.g. "A1:A10".
        width: Chart width in cm.
        height: Chart height in cm.
    """
    wb = _load_workbook(path)
    ws = _get_ws(wb, sheet_name)

    chart_classes = {"bar": BarChart, "line": LineChart, "pie": PieChart}
    chart_cls = chart_classes.get(chart_type.lower())
    if not chart_cls:
        wb.close()
        return f"Error: unsupported chart type '{chart_type}'. Use: bar, line, pie."

    chart = chart_cls()
    chart.title = title
    chart.width = width
    chart.height = height

    min_col, min_row, max_col, max_row = range_boundaries(data_range)
    data_ref = Reference(ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
    chart.add_data(data_ref, titles_from_data=True)

    if categories_range:
        cmin_col, cmin_row, cmax_col, cmax_row = range_boundaries(categories_range)
        cats = Reference(ws, min_col=cmin_col, min_row=cmin_row, max_col=cmax_col, max_row=cmax_row)
        chart.set_categories(cats)

    ws.add_chart(chart, target_cell)
    _save_workbook(wb, path)
    wb.close()
    return f"Created {chart_type} chart at {target_cell} in '{ws.title}' ({path})"
